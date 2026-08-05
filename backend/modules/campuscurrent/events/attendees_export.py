"""Print-ready CSV / XLSX export for event attendees."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from backend.modules.auth.models import User
from backend.modules.campuscurrent.models import Event

CAMPUS_TZ = ZoneInfo("Asia/Almaty")


def _format_going_at(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=ZoneInfo("UTC"))
    return value.astimezone(CAMPUS_TZ).strftime("%Y-%m-%d %H:%M")


def _format_event_when(event: Event) -> str:
    start = event.start_datetime
    if start.tzinfo is None:
        start = start.replace(tzinfo=ZoneInfo("UTC"))
    return start.astimezone(CAMPUS_TZ).strftime("%Y-%m-%d %H:%M")


def _full_name(user: User) -> str:
    return f"{user.name} {user.surname}".strip()


def build_attendees_csv(event: Event, rows: list[tuple[User, datetime]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["nuspace — Event attendance checklist"])
    writer.writerow([f"Event: {event.name}"])
    writer.writerow([f"Place: {event.place}"])
    writer.writerow([f"Starts: {_format_event_when(event)}"])
    writer.writerow([f"Total going: {len(rows)}"])
    writer.writerow([])
    writer.writerow(["#", "Full name", "Email", "Marked going at", "Arrived"])
    for index, (user, going_at) in enumerate(rows, start=1):
        writer.writerow(
            [
                index,
                _full_name(user),
                user.email,
                _format_going_at(going_at),
                "☐",
            ]
        )
    # Excel-friendly UTF-8 with BOM
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def build_attendees_xlsx(event: Event, rows: list[tuple[User, datetime]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    brand_font = Font(name="Calibri", size=16, bold=True, color="1E3A5F")
    meta_font = Font(name="Calibri", size=11, color="334155")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    alt_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws["A1"] = "nuspace"
    ws["A1"].font = brand_font
    ws.merge_cells("A1:E1")

    ws["A2"] = "Event attendance checklist — print and mark arrivals at the door"
    ws["A2"].font = meta_font
    ws.merge_cells("A2:E2")

    ws["A3"] = f"Event: {event.name}"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True)
    ws.merge_cells("A3:E3")

    ws["A4"] = f"Place: {event.place}  ·  Starts: {_format_event_when(event)}"
    ws["A4"].font = meta_font
    ws.merge_cells("A4:E4")

    ws["A5"] = f"Total going: {len(rows)}"
    ws["A5"].font = Font(name="Calibri", size=12, bold=True, color="1E3A5F")
    ws.merge_cells("A5:E5")

    headers = ["#", "Full name", "Email", "Marked going at", "Arrived"]
    header_row = 7
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for index, (user, going_at) in enumerate(rows, start=1):
        row_idx = header_row + index
        values = [
            index,
            _full_name(user),
            user.email,
            _format_going_at(going_at),
            "☐",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col in (1, 5) else "left",
            )
            if index % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 22

    ws.row_dimensions[header_row].height = 24
    widths = [6, 28, 34, 20, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A8"
    ws.print_title_rows = "1:7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True

    footer_row = header_row + len(rows) + 2
    ws.cell(
        row=footer_row,
        column=1,
        value="Generated by nuspace · Do not edit — mark Arrived on printout at check-in",
    ).font = Font(name="Calibri", size=9, italic=True, color="64748B")
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
